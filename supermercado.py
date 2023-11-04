import openpyxl


workbook = openpyxl.Workbook()


sheet = workbook.active
sheet.title = "Lista de Supermercados"


sheet["A1"] = "Supermercado"
sheet["B1"] = "Producto"
sheet["C1"] = "Precio"


supermercados = [
    ("Supermercado 1", "Limones", 5.33),
    ("Supermercado 1", "Salchichas", 7.40),
    ("Supermercado 2", "Bananos", 12.00),
    ("Supermercado 2", "Guayabas", 4.32),
    
]


for row, data in enumerate(supermercados, start=2):
    sheet.cell(row=row, column=1, value=data[0])
    sheet.cell(row=row, column=2, value=data[1])
    sheet.cell(row=row, column=3, value=data[2])


workbook.save("supermercado.xlsx")


workbook.close()
